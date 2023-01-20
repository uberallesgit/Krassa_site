import PySimpleGUI as sg
import selenium
import warnings
import csv, os
from time import sleep
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from datetime import datetime as dt
from datetime import timedelta as td
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook

working_directory =os.getcwd()
label_year = dt.today().strftime("%Y")
label_day = dt.today().strftime("%d")
label_month = dt.today().strftime("%m")


sg.theme('DarkAmber')

layout = [                        [sg.Text('                   Меню на сайт школы')],
            [sg.Text('Путь к Файлу'), sg.InputText(key="-FILE_PATH-"), sg.FileBrowse(button_text='Выбрать Файл',initial_folder=working_directory,file_types=[("XL Files","*.xls *.xlsx *.csv")])],
                                  [sg.Text('Дата'),sg.InputText(key="-FILE_DAY-",size=(2,1),default_text=label_day),sg.Text('.'),sg.InputText(key="-FILE_MONTH-",size=(2,1),default_text=label_month),sg.Text('.'),sg.Text(f"{label_year}",size=(4,1)),sg.Button('Мутация даты',key="-MUTATION-")],
                                  [sg.Text("",key="-NEW_FILE-")],
            [sg.Button('Отправить на сайт-->',key="-SUBMIT-"), sg.Button('Закрыть приложение',key="-CLOSE-"),sg.Button('Очистить все поля',key="-CLEAR-")]]





class FileSender:
    def time_now(self):
        self.today_is = dt.today()
        return self.today_is.strftime("%d.%m.%Y")
        print(time_now())



    def date_mutation(self):
        from openpyxl import load_workbook

        mutation_date = f"{label_year}.{values['-FILE_DAY-']}.{values['-FILE_MONTH-']}"

        file_name = values["-FILE_PATH-"]
        self.mutated_file = file_name.replace(file_name.split()[2],mutation_date+" new")
        print(self.mutated_file)
        wb = load_workbook(filename=f"{file_name}")
        ws = wb.active
        ws["J1"] = mutation_date
        wb.save(self.mutated_file)
        print(f"Дата мутировала на {mutation_date}")


    def run_driver(self):
        print("[INFO] Запускаем гекона")
        warnings.filterwarnings("ignore")  # отключение предупреждений об устаревшем стиле
        options = webdriver.FirefoxOptions()
        options.set_preference("general.useragent.override",
                               "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/101.0.4951.54 Safari/537.36")
        self.driver = webdriver.Firefox(executable_path="geckodriver.exe")

        print("Гекон запущен!")

    def open_school_page(self):
        if values["-FILE_PATH-"] == "":
            print("Файл не выбран!")
        else:
            print(values["-FILE_PATH-"])
            file_sender.run_driver()
            # self.driver.get("https://school-perv.educrimea.ru/")
            file_sender.school_login()
            file_sender.school_upload()

    def school_login(self):
        print("[INFO] Авторизуемся на школьном сайте")
        self.driver.get("https://school-perv.educrimea.ru/user/login")
        self.driver.implicitly_wait(20)
        login_field = self.driver.find_element(By.ID,"loginform-email")
        login_field.send_keys("uber0284@gmail.com")
        sleep(3)
        password_field = self.driver.find_element(By.ID,"loginform-password")
        password_field.send_keys("Reremedy1")
        sleep(3)
        enter_button = self.driver.find_element(By.NAME,"login-button")
        enter_button.click()
        sleep(3)
        print("[Авторизация удалась на славу!]")

    def school_upload(self):
        print("[INFO] Загружаем файл на школьный сайт")
        action = ActionChains(self.driver)
        sleep(3)
        conditions_link = self.driver.find_element(By.LINK_TEXT,"Условия")
        conditions_link.click()
        sleep(5)
        print("Вот сейчас")
        food_link = self.driver.find_element(By.LINK_TEXT,"Мониторинг горячего питания Минпросвещения РФ").click()
        sleep(7)
        edit_link = self.driver.find_element(By.LINK_TEXT,"Редактировать").click()
        sleep(7)
        add_files = self.driver.find_element(By.LINK_TEXT,"Добавить файлы").click()
        self.driver.switch_to.frame(4)
        sleep(3)
        activation = self.driver.find_element(By.ID,"nav-l1_Lw")
        action.move_to_element(activation).perform()
        sleep(3)
        upload_files = self.driver.find_element(By.XPATH,"/html/body/div/div[1]/div[4]/div[3]/span[1]")
        upload_files.click()
        sleep(5)
        hover = self.driver.find_element(By.CSS_SELECTOR,"div.ui-button:nth-child(3)")
        sleep(4)
        action.move_to_element(hover).perform()
        sleep(3)
        choose_file = self.driver.find_element(By.XPATH,"//input[@type='file']")
        choose_file.send_keys(rf"{file_path}")
        sleep(5)
        arrow_button = self.driver.find_element(By.CSS_SELECTOR,"div.elfinder-buttonset:nth-child(3) > div:nth-child(1)")
        action.move_to_element(arrow_button).perform()
        sleep(5)
        arrow_button = self.driver.find_element(By.CLASS_NAME,"ui-state-default.elfinder-button.ui-state-hover")
        arrow_button.click()
        sleep(5)
        self.driver.switch_to.default_content() #возврат  в начальный iFrame
        print("Загрузка определенно удалась!")
        # file_sender.choose_date() # метод для вызова календаря(ебёт мозги)

    def choose_date(self):
        print("[INFO] Меняем дату на актуальную.")
        self.day = values["-FILE_DAY-"]
        self.month = values["-FILE_MONTH-"]
        if self.day[0] =="0":
            self.day = self.day.replace("0","")
        if self.month[0] =="0":
            self.month = self.month.replace("0","")
            # тут(найти по тегу)
        print("[INFO]Поиск указанного дня...")
        action = ActionChains(self.driver)
        calendar_call = self.driver.find_element(By.PARTIAL_LINK_TEXT,f"{self.day}").click()
        print(calendar_call)
        sleep(3)
        pick_a_day = self.driver.find_element(By.XPATH,f"//td[contains(text(),'{self.day}')]")

#############################################################################################################################
        # https://stackoverflow.com/questions/44777053/selenium-movetargetoutofboundsexception-with-firefox
        # from selenium import webdriver
        # from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
        # from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
        # from selenium.webdriver.common.action_chains import ActionChains
        #
        # binary = FirefoxBinary('C:\\Program Files\\Mozilla Firefox\\firefox.exe')
        # caps = DesiredCapabilities().FIREFOX
        # caps["marionette"] = True
        # driver = webdriver.Firefox(capabilities=caps, firefox_binary=binary,
        #                            executable_path="C:\\Utility\\BrowserDrivers\\geckodriver.exe")
        # driver.get("https://stackoverflow.com")
        # last_height = driver.execute_script("return document.body.scrollHeight")
        # driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        # source_element = driver.find_element_by_xpath('//*[@id="footer"]/div/ul/li[1]/a')
        # ActionChains(driver).move_to_element(source_element).perform()
###################################################################################################################################


        action.move_to_element(pick_a_day).perform()


        # pick_a_day = WebDriverWait(self.driver,30).until(EC.element_to_be_clickable((By.XPATH,f"//td[contains(text(),'11')]")))
        print("Календарь найден!Дата установлена по феншую!")
        sleep(1)

        print("[INFO] Файл загружен на школьный сайт!")
        print("[INFO] Для проверки перейдите по ссылке: https://school-perv.educrimea.ru/food")

    def clear_all(self):
        window["-FILE_PATH-"].update("")
        window["-FILE_DAY-"].update("")
        window["-FILE_MONTH-"].update("")




window = sg.Window('Krassa 1.0', layout)

file_sender = FileSender()

# Цикл событий (Event-Loop)

while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event == '-CLOSE-':
        break

    elif event =="-SUBMIT-":

        file_path = values["-FILE_PATH-"].replace("/","\\") # для винды (для линукса replace не нужен )

        file_sender.open_school_page()
    elif event == "-MUTATION-":
        file_sender.date_mutation()
        window["-NEW_FILE-"].update("Дата изменена, файл переименован !")
    elif event == "-CLEAR-":
        file_sender.clear_all()



window.close()
#datetime lin